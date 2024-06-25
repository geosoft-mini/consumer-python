import pandas as pd
from openpyxl import load_workbook


 
# Generating workbook and writer engine
wb = load_workbook("apk.xlsx")

ws = wb.create_sheet('test')

wb.save('./apk.xlsx')
wb.close()

